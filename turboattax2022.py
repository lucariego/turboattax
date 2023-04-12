import openpyxl

workbook = openpyxl.load_workbook('trading_cards.xlsx')
worksheet = workbook.active

mode = 0

def addCards():
    currentCard = ""
    while currentCard != "e":
        currentCard = input("Enter card number: ")
        if currentCard == "e":
            break

        adder(currentCard)            
            
def adder(currentCard):
    rows = list(worksheet.iter_rows(values_only=True))
    quantity = rows[int(currentCard)][1] + 1
    print(quantity)
    worksheet.cell(row = int(currentCard) + 1, column = 2, value = quantity )
    workbook.save('trading_cards.xlsx')
    print(f"Card added. You have {quantity} units of card {currentCard}.")

def checkCards():
    currentCard = ""
    while currentCard != "e":
        currentCard = input("Enter card number: ")
        if currentCard == "e":
            break

        rows = list(worksheet.iter_rows(values_only=True))
        quantity = rows[int(currentCard)][1]
        print(f"There are {quantity} units of card {currentCard}.")

def checkWhileAdd():
    currentCard = ""
    while currentCard != "e":
        currentCard = input("Enter card number: ")
        if currentCard == "e":
            break

        rows = list(worksheet.iter_rows(values_only=True))
        quantity = rows[int(currentCard)][1]
        print(f"There were {quantity} units of card {currentCard}.")
        adder(currentCard)

def missingCards():
    for row in worksheet.iter_rows(values_only=True):
        if row[1] == 0:
            print(f"Card {row[0] } is missing.")

def verify():
    count = 3
    password = ""
    while (count > 0):
        password = input("Enter password: ")
        if password != "goatifi":
            print("\nIncorrect password. You have %s attempts left." %(count - 1))
            count -= 1
        else:
            return True
    


def eraseExcel():
    count = 1
    while (count < 382):
            worksheet.cell(row = count + 1, column = 2, value = 0)
            count += 1
    workbook.save('trading_cards.xlsx')
    

while mode != "e":
    print("\nWelcome to Turbo Attax 2022!. Choose a mode:")
    print("1. Check cards")
    print("2. Add cards")
    print("3. Missing cards")
    print("4. Check while adding cards")
    #print("5. Erase Excel (CAREFUL)")
    print("Type 'e' to exit.")

    mode = input("\nEnter mode: ")

    if mode == "1":
        checkCards()
    elif mode == "2":
        addCards()
    elif mode == "3":
        missingCards()
    elif mode == "4":
        checkWhileAdd()
    elif mode == "5":
        if verify():
            eraseExcel()
            print("\nExcel file erased.")
    else:
        print("Invalid mode. Please try again.")
